import openpyxl
import re
from datetime import datetime, timedelta, date

def get_excel_date(serial):
    try:
        val = float(serial)
        if val >= 60:
            val -= 1
        dt = datetime(1899, 12, 30) + timedelta(days=val)
        return dt.strftime('%Y-%m-%d')
    except:
        return str(serial)

def parse_date_cell(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        val_str = val.strip()
        match = re.search(r'(\d+)[-/](\d+)[-/](\d+)', val_str)
        if match:
            d, m, y = map(int, match.groups())
            if y < 100:
                y += 2000
            try:
                return datetime(y, m, d).strftime('%Y-%m-%d')
            except:
                pass
    if isinstance(val, (int, float)) and 40000 <= val <= 50000:
        return get_excel_date(val)
    return None

def evaluate_cell(ws, cell, visited=None):
    if visited is None:
        visited = set()
        
    cell_coord = cell.coordinate
    if cell_coord in visited:
        return 0
    visited.add(cell_coord)
    
    val = cell.value
    if val is None:
        return None
    if not isinstance(val, str) or not val.startswith('='):
        return val
        
    formula = val[1:].strip()
    
    # Check for SUM
    sum_match = re.match(r'^SUM\((.+)\)$', formula, re.IGNORECASE)
    if sum_match:
        cell_range = sum_match.group(1)
        total = 0
        for area in cell_range.split(','):
            area = area.strip()
            if ':' in area:
                try:
                    for row_cells in ws[area]:
                        for c in row_cells:
                            c_val = evaluate_cell(ws, c, visited.copy())
                            if isinstance(c_val, (int, float)):
                                total += c_val
                except:
                    pass
            else:
                try:
                    c_val = evaluate_cell(ws, ws[area], visited.copy())
                    if isinstance(c_val, (int, float)):
                        total += c_val
                except:
                    pass
        return total
        
    # General expression
    formula_clean = formula.replace('$', '')
    
    # Find cell references
    refs = re.findall(r'\b([A-Z]+\d+)\b', formula_clean)
    for ref in set(refs):
        try:
            ref_val = evaluate_cell(ws, ws[ref], visited.copy())
            if ref_val is None:
                ref_val = 0
            formula_clean = re.sub(r'\b' + ref + r'\b', str(ref_val), formula_clean)
        except:
            pass
            
    formula_clean = formula_clean.replace('--', '+')
    
    if re.match(r'^[0-9\.\+\-\*/\(\)\s]+$', formula_clean):
        try:
            return eval(formula_clean, {"__builtins__": None}, {})
        except:
            return None
    else:
        return val

def parse_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # Financial target is in sheet 1 cell AU1 or AS1?
    fi_target = 4500000000.0
    for sname in wb.sheetnames:
        ws = wb[sname]
        for cell_ref in ['AW1', 'AU1', 'AS1']:
            val = evaluate_cell(ws, ws[cell_ref])
            if val and isinstance(val, (int, float)) and val > 1000000000:
                fi_target = float(val)
                break
                
    all_transactions = []
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        
        # Default start date based on Sheet name or row 4 cell A
        start_month_str = evaluate_cell(ws, ws.cell(row=4, column=1))
        current_date_base = None
        if start_month_str:
            match = re.search(r'THÁNG (\d+)\s+NĂM\s+(\d+)', str(start_month_str), re.IGNORECASE)
            if match:
                month = int(match.group(1))
                year = int(match.group(2))
                current_date_base = datetime(year, month, 1)
                
        if current_date_base is None:
            current_date_base = datetime(2025, 9, 1)
            
        current_income_date = current_date_base.strftime('%Y-%m-%d')
        current_expense_date = current_date_base.strftime('%Y-%m-%d')
        
        # Find the summation row (starts with + in column B or contains +)
        end_row = ws.max_row
        for r in range(4, ws.max_row + 1):
            b_val = evaluate_cell(ws, ws.cell(row=r, column=2))
            if b_val in ['+', '%'] or str(b_val).strip() == '+':
                end_row = r - 1
                break
                
        # Iterate over rows
        for r in range(4, end_row + 1):
            col_s_val = evaluate_cell(ws, ws.cell(row=r, column=openpyxl.utils.column_index_from_string('S')))
            col_c_val = evaluate_cell(ws, ws.cell(row=r, column=openpyxl.utils.column_index_from_string('C')))
            col_d_val = evaluate_cell(ws, ws.cell(row=r, column=openpyxl.utils.column_index_from_string('D')))
            
            # Check for date in Column C (for income section)
            c_date = parse_date_cell(col_c_val)
            if c_date is not None:
                current_income_date = c_date
                
            # Check for date in Column S (for expense/debt section)
            s_date = parse_date_cell(col_s_val)
            if s_date is not None:
                current_expense_date = s_date
            
            # Special case for Lương: val in Q directly
            salary_val = evaluate_cell(ws, ws.cell(row=r, column=openpyxl.utils.column_index_from_string('Q')))
            if salary_val and isinstance(salary_val, (int, float)):
                all_transactions.append({
                    'sheet': sname,
                    'row': r,
                    'date': current_income_date,
                    'group': 'KHOẢN THU',
                    'category': 'Lương',
                    'item': 'Lương tháng',
                    'amount': float(salary_val)
                })
                
            mappings = [
                ('C', 'D', 'KHOẢN THU', 'Buôn bán'),
                ('E', 'F', 'KHOẢN THU', 'Thu phí'),
                ('G', 'H', 'KHOẢN THU', 'Nhận tiền D'),
                ('I', 'J', 'KHOẢN THU', 'Mượn'),
                ('K', 'L', 'KHOẢN THU', 'Nợ trả'),
                ('M', 'N', 'KHOẢN THU', 'Quỹ'),
                ('O', 'P', 'KHOẢN THU', 'Có sẵn / Khác'),
                ('S', 'T', 'KHOẢN CHI', 'Nhu yếu phẩm'),
                ('U', 'V', 'KHOẢN CHI', 'Ăn uống'),
                ('W', 'X', 'KHOẢN CHI', 'Mua sắm'),
                ('Y', 'Z', 'KHOẢN CHI', 'Thư giãn'),
                ('AA', 'AB', 'KHOẢN CHI', 'Phát sinh'),
                ('AC', 'AD', 'KHOẢN CHI', 'Cấp tiền TG'),
                ('AE', 'AF', 'KHOẢN CHI', 'Mua đồ d'),
                ('AI', 'AH', 'KHOẢN NỢ', 'Khoản nợ'),
                ('AL', 'AK', 'ĐÒI NỢ', 'Đòi nợ'),
                ('AN', 'AO', 'DOANH THU ĐẦU TƯ', 'Doanh thu đầu tư')
            ]
            
            for name_col, val_col, group, subcat in mappings:
                n_idx = openpyxl.utils.column_index_from_string(name_col)
                v_idx = openpyxl.utils.column_index_from_string(val_col)
                name_val = evaluate_cell(ws, ws.cell(row=r, column=n_idx))
                val_val = evaluate_cell(ws, ws.cell(row=r, column=v_idx))
                
                if name_val is not None and parse_date_cell(name_val) is not None:
                    continue
                
                if val_val is not None and isinstance(val_val, (int, float)) and val_val != 0:
                    item_name = str(name_val) if name_val is not None else subcat
                    tx_date = current_income_date if group == 'KHOẢN THU' else current_expense_date
                    all_transactions.append({
                        'sheet': sname,
                        'row': r,
                        'date': tx_date,
                        'group': group,
                        'category': subcat,
                        'item': item_name.strip(),
                        'amount': float(val_val)
                    })
                    
    return fi_target, all_transactions

def insert_rows_and_shift_merged(ws, row_idx, amount):
    ws.insert_rows(row_idx, amount)
    # Shift merged cells manually due to openpyxl insert_rows bug
    for r in list(ws.merged_cells.ranges):
        if r.min_row >= row_idx:
            ws.merged_cells.remove(r)
            r.shift(row_shift=amount, col_shift=0)
            ws.merged_cells.add(r)

def add_transaction(file_path, tx):
    """
    tx should be a dict: {
        'date': 'YYYY-MM-DD',
        'group': 'KHOẢN THU' or 'KHOẢN CHI' or 'KHOẢN NỢ' or 'ĐÒI NỢ' or 'DOANH THU ĐẦU TƯ',
        'category': 'Buôn bán', 'Ăn uống', etc.,
        'item': 'Description',
        'amount': 150000.0
    }
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sname = wb.sheetnames[-1]
    ws = wb[sname]
    
    # Find summation row
    plus_row = None
    for r in range(4, ws.max_row + 1):
        b_val = evaluate_cell(ws, ws.cell(row=r, column=2))
        if b_val == '+' or str(b_val).strip() == '+':
            plus_row = r
            break
            
    if plus_row is None:
        raise ValueError("Could not find summation row (+) in sheet")
        
    print(f"Adding transaction to sheet '{sname}' before row {plus_row}")
    
    mappings = {
        'Buôn bán': ('C', 'D'),
        'Thu phí': ('E', 'F'),
        'Nhận tiền D': ('G', 'H'),
        'Mượn': ('I', 'J'),
        'Nợ trả': ('K', 'L'),
        'Quỹ': ('M', 'N'),
        'Có sẵn / Khác': ('O', 'P'),
        'Lương': (None, 'Q'),
        'Nhu yếu phẩm': ('S', 'T'),
        'Ăn uống': ('U', 'V'),
        'Mua sắm': ('W', 'X'),
        'Thư giãn': ('Y', 'Z'),
        'Phát sinh': ('AA', 'AB'),
        'Cấp tiền TG': ('AC', 'AD'),
        'Mua đồ d': ('AE', 'AF'),
        'Khoản nợ': ('AI', 'AH'),
        'Đòi nợ': ('AL', 'AK'),
        'Doanh thu đầu tư': ('AN', 'AO')
    }
    
    cat = tx['category']
    if cat not in mappings:
        raise ValueError(f"Unknown category: {cat}")
        
    name_col, val_col = mappings[cat]
    
    # Check last date in the appropriate column (C for Income, S for others)
    date_col_letter = 'C' if tx['group'] == 'KHOẢN THU' else 'S'
    date_col_idx = openpyxl.utils.column_index_from_string(date_col_letter)
    
    last_date_str = None
    for r in range(plus_row - 1, 3, -1):
        d_val = evaluate_cell(ws, ws.cell(row=r, column=date_col_idx))
        parsed_d = parse_date_cell(d_val)
        if parsed_d is not None:
            last_date_str = parsed_d
            break
            
    tx_dt = datetime.strptime(tx['date'], '%Y-%m-%d')
    tx_date_str = f"Ngày {tx_dt.strftime('%d/%m/%Y')}"
    
    need_date_header = (last_date_str != tx['date'])
            
    rows_inserted = 0
    if need_date_header:
        insert_rows_and_shift_merged(ws, plus_row, 1)
        date_row = plus_row
        
        ws.cell(row=date_row, column=date_col_idx, value=tx_date_str)
        
        # Merge cells for clean appearance
        if tx['group'] == 'KHOẢN THU':
            p_col_idx = openpyxl.utils.column_index_from_string('P')
            ws.merge_cells(start_row=date_row, start_column=date_col_idx, end_row=date_row, end_column=p_col_idx)
        else:
            af_col_idx = openpyxl.utils.column_index_from_string('AF')
            ws.merge_cells(start_row=date_row, start_column=date_col_idx, end_row=date_row, end_column=af_col_idx)
            
        font_date = openpyxl.styles.Font(name="Arial", size=10, bold=True, color="1E293B")
        align_date = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws.cell(row=date_row, column=date_col_idx).font = font_date
        ws.cell(row=date_row, column=date_col_idx).alignment = align_date
        
        plus_row += 1
        rows_inserted += 1
        
    insert_rows_and_shift_merged(ws, plus_row, 1)
    tx_row = plus_row
    
    # Index TT
    last_idx = 1
    for r in range(tx_row - 1, 3, -1):
        b_val = evaluate_cell(ws, ws.cell(row=r, column=2))
        if b_val and str(b_val).isdigit():
            last_idx = int(b_val)
            break
            
    new_idx = last_idx + 1
    ws.cell(row=tx_row, column=2, value=new_idx)
    ws.cell(row=tx_row, column=openpyxl.utils.column_index_from_string('R'), value=f"=B{tx_row}")
    
    if name_col:
        n_idx = openpyxl.utils.column_index_from_string(name_col)
        ws.cell(row=tx_row, column=n_idx, value=tx['item'])
        
    v_idx = openpyxl.utils.column_index_from_string(val_col)
    ws.cell(row=tx_row, column=v_idx, value=float(tx['amount']))
    
    ws.cell(row=tx_row, column=v_idx).number_format = '#,##0'
    ws.cell(row=tx_row, column=v_idx).alignment = openpyxl.styles.Alignment(horizontal="right")
    
    font_data = openpyxl.styles.Font(name="Arial", size=10)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=tx_row, column=col)
        if cell.value is not None:
            if not cell.font or cell.font.name != "Arial":
                cell.font = font_data
                
    plus_row += 1
    rows_inserted += 1
    
    # Update summaries
    new_sum_limit = plus_row - 1
    sum_cols = {
        'C': f"=SUM(D4:D{new_sum_limit})",
        'E': f"=SUM(F4:F{new_sum_limit})",
        'G': f"=SUM(H4:H{new_sum_limit})",
        'I': f"=SUM(J4:J{new_sum_limit})",
        'K': f"=SUM(L4:L{new_sum_limit})",
        'M': f"=SUM(N4:N{new_sum_limit})",
        'O': f"=SUM(P4:P{new_sum_limit})",
        'Q': f"=SUM(C{plus_row}:P{plus_row})+SUM(Q4:Q{new_sum_limit})",
        'S': f"=SUM(T4:T{new_sum_limit})",
        'U': f"=SUM(V4:V{new_sum_limit})",
        'W': f"=SUM(X4:X{new_sum_limit})",
        'Y': f"=SUM(Z4:Z{new_sum_limit})",
        'AA': f"=SUM(AB4:AB{new_sum_limit})",
        'AC': f"=SUM(AD4:AD{new_sum_limit})",
        'AE': f"=SUM(AF4:AF{new_sum_limit})",
        'AG': f"=SUM(S{plus_row}:AF{plus_row})",
        'AH': f"=SUM(AH4:AH{new_sum_limit})",
        'AK': f"=SUM(AK4:AK{new_sum_limit})",
        'AN': f"=SUM(AO4:AO{new_sum_limit})"
    }
    
    for col_letter, formula in sum_cols.items():
        c_idx = openpyxl.utils.column_index_from_string(col_letter)
        ws.cell(row=plus_row, column=c_idx, value=formula)
        ws.cell(row=plus_row, column=c_idx).number_format = '#,##0'
        
    # Percentages
    pct_row = plus_row + 1
    pct_cols = {
        'C': f"=(C{plus_row}/$Q${plus_row})*100",
        'E': f"=(E{plus_row}/$Q${plus_row})*100",
        'G': f"=(G{plus_row}/$Q${plus_row})*100",
        'I': f"=(I{plus_row}/$Q${plus_row})*100",
        'K': f"=(K{plus_row}/$Q${plus_row})*100",
        'M': f"=(M{plus_row}/$Q${plus_row})*100",
        'O': f"=(O{plus_row}/$Q${plus_row})*100",
        'S': f"=(S{plus_row}/$AG${plus_row})*100",
        'U': f"=(U{plus_row}/$AG${plus_row})*100",
        'W': f"=(W{plus_row}/$AG${plus_row})*100",
        'Y': f"=(Y{plus_row}/$AG${plus_row})*100",
        'AA': f"=(AA{plus_row}/$AG${plus_row})*100",
        'AC': f"=(AC{plus_row}/$AG${plus_row})*100",
        'AE': f"=(AE{plus_row}/$AG${plus_row})*100"
    }
    
    for col_letter, formula in pct_cols.items():
        c_idx = openpyxl.utils.column_index_from_string(col_letter)
        ws.cell(row=pct_row, column=c_idx, value=formula)
        ws.cell(row=pct_row, column=c_idx).number_format = '0.00'
        
    wb.save(file_path)
    print(f"Successfully added transaction and updated formulas. Rows inserted: {rows_inserted}")
